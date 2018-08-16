from urllib import request
from urllib import parse
import re
import http.cookiejar
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook
import sys
from openpyxl.styles import Font, colors, Alignment
from tkinter.filedialog import *
import getpass
class URP:

    def __init__(self,username,password):

        self.usr = username
        self.psw = password
        self.all_row =[]
        self.hosturl= 'http://202.197.144.243:8083'
        self.posturl =  'http://202.197.144.243:8083/loginAction.do'
        try:

            cj = http.cookiejar.LWPCookieJar()
            cookie_support = request.HTTPCookieProcessor(cj)
            openr = request.build_opener(cookie_support)
            request.install_opener(openr)

            request.urlopen(self.hosturl)
        except Exception:
            messagebox.showerror('提示', '无法访问')
            sys.exit(0)

        #请求头
        self.headers  ={
            'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.119 Safari/537.36',
            'Referer': 'http://202.197.144.243:8083/'
            }
        #提交的数据
        self.postData ={
            'zjh':self.usr,
            'mm':self.psw
        }
        #更改数据的编码
        self.postData = parse.urlencode(self.postData).encode('GBK')
        #放到req
        self.req = request.Request(self.posturl,self.postData,self.headers)

    def login(self):
        try :
            #进入主页面
           request.urlopen(self.req)
           self.resp = request.urlopen('http://202.197.144.243:8083/xkAction.do?actionType=6')
           self.html =self.resp.read()
           self.html = self.html.decode('GBK')

           return self.html

        except Exception as e:
            messagebox.showerror('提示','%s无法访问："http://202.197.144.243:8083/"'%e.code)





    def get_Table(self,html):
        r1 = r'<table cellpadding="0" width="100%" class="displayTag" cellspacing="1" border="0" id="user">([\s\S]*?)</table>'
        r2 = r'<td[\s\S]*?>([\S\s]*?)</td>'
        table_list = re.findall(r1,html)
        td_list=re.findall(r2,table_list[0])
        for i in  range(len(td_list)):
               td_list[i]= ''.join(td_list[i]).replace('\r','')
               td_list[i]= ''.join(td_list[i]).replace('\n','')
               td_list[i] =''.join(td_list[i]).replace('\t','')
               td_list[i] =''.join(td_list[i]).replace('&nbsp;','')
               td_list[i] =''.join(td_list[i]).replace('<br>','')
               td_list[i] =''.join(td_list[i]).replace('<div align="center">','')
               td_list[i] =''.join(td_list[i]).replace('</div>','')
               td_list[i] =''.join(td_list[i]).replace('<p class="style4">','')
               td_list[i] = ''.join(td_list[i]).replace('</p>', '')
               td_list[i] = ''.join(td_list[i]).replace('<p align="center" class="td2 style5"><strong>', '')
               td_list[i] = ''.join(td_list[i]).replace('</strong>', '').strip()

        return td_list

    def write_xlsx(self,path,data):
        try:
            bold = Font(name='等线', size=10, italic=False, color=colors.BLACK, bold=True)
            #定义表格
            wb = Workbook()
            ws = wb.active
            ws.title='课表'
            ws.merge_cells('A1:B1')
            ws.merge_cells('A6:I6')
            ws.merge_cells('A11:I11')
            ws.merge_cells('A2:A5')
            ws.merge_cells('A7:A10')
            ws.merge_cells('A12:A13')
            ws.row_dimensions[1].height = 50
            ws.row_dimensions[2].height = 35
            ws.row_dimensions[3].height = 35
            ws.row_dimensions[4].height = 35
            ws.row_dimensions[5].height = 35
            ws.row_dimensions[6].height = 20
            ws.row_dimensions[7].height = 35
            ws.row_dimensions[8].height = 35
            ws.row_dimensions[9].height = 35
            ws.row_dimensions[11].height = 20
            ws.row_dimensions[10].height = 35
            ws.row_dimensions[12].height = 35
            ws.row_dimensions[13].height = 35
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 45
            ws.column_dimensions['D'].width = 45
            ws.column_dimensions['E'].width = 45
            ws.column_dimensions['F'].width = 45
            ws.column_dimensions['G'].width = 45
            ws.column_dimensions['H'].width = 45
            ws.column_dimensions['I'].width = 45
            #写入数据
            row1 =['A1','C1','D1','E1','F1','G1','H1','I1']
            row2 =['A2','B2','C2','D2','E2','F2','G2','H2','I2']
            row3 =['B3','C3','D3','E3','F3','G3','H3','I3']
            row4 =['B4', 'C4', 'D4', 'E4', 'F4', 'G4', 'H4', 'I4']
            row5 =['B5','C5','D5','E5','F5','G5','H5','I5']
            row6 = ['A6']
            row7 =['A7','B7','C7','D7','E7','F7','G7','H7','I7']
            row8 = [ 'B8', 'C8', 'D8', 'E8', 'F8', 'G8', 'H8', 'I8']
            row9 = ['B9', 'C9', 'D9', 'E9', 'F9', 'G9', 'H9', 'I9']
            row10 = ['B10', 'C10', 'D10', 'E10', 'F10', 'G10', 'H10', 'I10']
            row11 = ['A11']
            row12 = ['A12', 'B12', 'C12', 'D12', 'E12', 'F12', 'G12', 'H12', 'I12']
            row13 = ['B13', 'C13', 'D13', 'E13', 'F13', 'G13', 'H13', 'I13']
            all_rows =row1+row2+row3+row4+row5+row6+row7+row8+row9+row10+row11+row12+row13
            for i in range(len(data)):
                ws[all_rows[i]] = data[i]
                ws[all_rows[i]].font=bold
                ws[all_rows[i]].alignment=Alignment(horizontal='center',vertical='center')
            wb.save(path)

        except Exception:
            messagebox.showerror('提示','文件已打开')
    def getName(self):
        request.urlopen('http://202.197.144.243:8083/xjInfoAction.do?oper=ydxx')
        self.resp = request.urlopen('http://202.197.144.243:8083/userInfo.jsp').read()
        self.html = self.resp.decode('GBK')
        return  self.html
    def info(self,html):
        #用户名匹配
        r1 = r'<td >([\s\S]*?)</td>'
        #匹配表格，这里会用到第二个表格
        r2 = r'<table width="100%" border="0" cellpadding="0" cellspacing="0" class="titleTop3">([\s\S]*?)</table>'
        #匹配表格的电话
        r3 = r'<td >[\s\S]*?<input type="text" name="dh" value="([\s\S]*?)" tag="电话|0|s|40">[\s\S]*?</td>'
        #匹配表格的地址
        r4 = r'<td>[\s\S]*?<input type="text" name="txdz" value="([\s\S]*?)" tag="通讯地址|0|s|100">[\s\S]*?</td>'
        #匹配表格的Email
        r5 = r'<td>[\s\S]*?<input type="text" name="email" value="([\s\S]*?)" tag="电子邮件|0|e|100">[\s\S]*?</td>'

        #找到的用户名
        name = re.findall(r1,html)
        #找到的所有表
        table =re.findall(r2,html)
        #找到的电话，这里匹配貌似有bug 。。。不管了
        phone = re.findall(r3,table[1])
        for i in phone:
            if i=='':
                phone.remove(i)
        #找到的通讯地址
        address = re.findall(r4,table[1])
        for i in address:
            if i=='':
                address.remove(i)
        email = re.findall(r5,table[1])
        for j in email:
            if j =='':
                email.remove(j)

        #放到字典里面
        info_dict={
            'name': ''.join(name[0]).strip(),
            'phone': ''.join(phone[0]).strip(),
            'address':''.join(address[0]).strip(),
            'email':''.join(email[0]).strip()
        }
        return info_dict

def query(usr,psw,option):
    if option ==0:
        urp= URP(usr,psw)
        html = urp.login()

        if html is not None:
            try:
                html1 = urp.getName()
                info =urp.info(html1)
                td_list = urp.get_Table(html)
                messagebox.showinfo('提示', '登录成功！  你好，' + info.get('name'))

                back = messagebox.askyesno('提取', '登录成功！\n是否提取课表？')
                if back == True:
                    if sys.platform =='win32':
                        savepath ='C:\\Users\\'+getpass.getuser()+'\\Desktop'
                    else:
                        savepath ='\\'
                    path = asksaveasfilename(title='课表',initialfile='课表.xlsx',initialdir=savepath,defaultextension='.xslx')
                    if path is not '':
                        urp.write_xlsx(path, td_list)
                        messagebox.showinfo('提示','提取成功！')
                else:
                    var.set('')
                    var2.set('')
                return True
            except request.HTTPError:
                messagebox.showerror('提示', '登录失败，账号或密码错误')
    else:
        urp = URP(usr, psw)
        html = urp.login()

        if html is not None:
            try:
                html1 = urp.getName()
                info = urp.info(html1)
                td_list = urp.get_Table(html)
                return True
            except request.HTTPError:
                pass




if __name__ == '__main__':


    def set():

        usr =e1.get()
        psw =e2.get()
        if usr==''or psw=='':
            messagebox.showerror('提示','用户名或密码为空')
        else:
            try:
                usr = int(usr)
            except Exception:
                messagebox.showwarning('提示','账号完全为数字才可以哟!')
            back = query(usr,psw,0)


    def exit():
          usr = e1.get()
          psw = e2.get()
          if (usr is '' or psw is ''):
              pass
          else:
              back = query(usr, psw,1)

              if back ==True:
                try:
                    request.urlopen('http://202.197.144.243:8083/logout.do').read()
                    messagebox.showinfo('提示', '账号已退出')
                except Exception:
                    messagebox.showerror('提示','无法访问')
                    sys.exit(0)
                var.set('')
                var2.set('')
              else:
                  pass



    root = Tk()
    root.config(bg='#808080')
    root.title('一键提取课表')
    root.geometry('300x170+650+300')
    root.resizable(0, 0)
    fr = Frame(root,bg='gray')
    fr.pack(fill=X)
    l1 = Label(fr,text='用户名',bg='gray',fg='yellow',height=2)
    l1.grid(row=0,column=0)
    var=StringVar()
    var2 = StringVar()
    e1 = Entry(fr,width='15',textvariable=var)
    e1.grid(row=0,column=1)
    l2 = Label(fr, text='密    码',bg='gray',fg='yellow',height=2)
    l2.grid(row=1, column=0)
    e2 = Entry(fr,width='15',textvariable=var2)
    e2['show'] = '*'
    e2.grid(row=1, column=1)
    btn1 =Button(fr,width=10,height=1,text='登录',bg='green',fg='white',command =set)
    btn1.grid(row =0,column =4)
    Label(fr,width=5,bg='gray').grid(row=1,column=3)
    btn2 =Button(fr,width=10,height=1,text='注销',bg='red',fg='white',command =exit)
    btn2.grid(row =1,column=4)
    Label(root,fg='blue',height=10,text='本人承诺，本程序不收集任何信息\n仅适用湖北第二师范学院(URP综合教务系统)',bg='gray').pack()
    root.mainloop()

