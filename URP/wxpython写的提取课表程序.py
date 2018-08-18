import wx
from urllib import request
from urllib import parse
import re
import http.cookiejar
import os
from openpyxl.styles import Font, colors, Alignment
from openpyxl import Workbook
import getpass
import sys
class URP:

    def __init__(self,usr,psw,port):

        self.usr = usr
        self.psw = psw
        self.port =port
        self.hosturl= 'http://202.197.144.243:'+port
        self.posturl='http://202.197.144.243:'+port+'/loginAction.do'
        cj = http.cookiejar.LWPCookieJar()
        cookie_support = request.HTTPCookieProcessor(cj)
        openr = request.build_opener(cookie_support)
        request.install_opener(openr)
        request.urlopen(self.hosturl)

        #请求头
        self.headers  ={
            'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.119 Safari/537.36',
            'Referer': 'http://202.197.144.243:'+port+'/'
            }
        #提交的数据
        self.postData ={
            'zjh':self.usr,
            'mm':self.psw
        }
        #更改数据的编码
        self.postData = parse.urlencode(self.postData).encode('GBK')
        #放到req
        self.req = request.Request(self.posturl,self.postData,self.headers)

    def login(self):
        try :
            #进入主页面
            request.urlopen(self.req)
            self.resp = request.urlopen('http://202.197.144.243:'+self.port+'/xkAction.do?actionType=6')
            self.html = self.resp.read()
            self.html = self.html.decode('GBK')
            return self.html
        except Exception:
            pass

    def getTable(self,html):
        r1 = r'<table cellpadding="0" width="100%" class="displayTag" cellspacing="1" border="0" id="user">([\s\S]*?)</table>'
        r2 = r'<td[\s\S]*?>([\S\s]*?)</td>'
        table_list = re.findall(r1, html)
        td_list = re.findall(r2, table_list[0])
        for i in range(len(td_list)):
            td_list[i] = ''.join(td_list[i]).replace('\r', '')
            td_list[i] = ''.join(td_list[i]).replace('\n', '')
            td_list[i] = ''.join(td_list[i]).replace('\t', '')
            td_list[i] = ''.join(td_list[i]).replace('&nbsp;', '')
            td_list[i] = ''.join(td_list[i]).replace('<br>', '')
            td_list[i] = ''.join(td_list[i]).replace('<div align="center">', '')
            td_list[i] = ''.join(td_list[i]).replace('</div>', '')
            td_list[i] = ''.join(td_list[i]).replace('<p class="style4">', '')
            td_list[i] = ''.join(td_list[i]).replace('</p>', '')
            td_list[i] = ''.join(td_list[i]).replace('<p align="center" class="td2 style5"><strong>', '')
            td_list[i] = ''.join(td_list[i]).replace('</strong>', '').strip()

        return td_list

    def write_xlsx(self, path, data):
        try:
            bold = Font(name='等线', size=10, italic=False, color=colors.BLACK, bold=True)
            # 定义表格
            wb = Workbook()
            ws = wb.active
            ws.title = '课表'
            ws.merge_cells('A1:B1')
            ws.merge_cells('A6:I6')
            ws.merge_cells('A11:I11')
            ws.merge_cells('A2:A5')
            ws.merge_cells('A7:A10')
            ws.merge_cells('A12:A13')
            ws.row_dimensions[1].height = 50
            ws.row_dimensions[2].height = 35
            ws.row_dimensions[3].height = 35
            ws.row_dimensions[4].height = 35
            ws.row_dimensions[5].height = 35
            ws.row_dimensions[6].height = 20
            ws.row_dimensions[7].height = 35
            ws.row_dimensions[8].height = 35
            ws.row_dimensions[9].height = 35
            ws.row_dimensions[11].height = 20
            ws.row_dimensions[10].height = 35
            ws.row_dimensions[12].height = 35
            ws.row_dimensions[13].height = 35
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 45
            ws.column_dimensions['D'].width = 45
            ws.column_dimensions['E'].width = 45
            ws.column_dimensions['F'].width = 45
            ws.column_dimensions['G'].width = 45
            ws.column_dimensions['H'].width = 45
            ws.column_dimensions['I'].width = 45
            # 写入数据
            row1 = ['A1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1']
            row2 = ['A2', 'B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'I2']
            row3 = ['B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3', 'I3']
            row4 = ['B4', 'C4', 'D4', 'E4', 'F4', 'G4', 'H4', 'I4']
            row5 = ['B5', 'C5', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5']
            row6 = ['A6']
            row7 = ['A7', 'B7', 'C7', 'D7', 'E7', 'F7', 'G7', 'H7', 'I7']
            row8 = ['B8', 'C8', 'D8', 'E8', 'F8', 'G8', 'H8', 'I8']
            row9 = ['B9', 'C9', 'D9', 'E9', 'F9', 'G9', 'H9', 'I9']
            row10 = ['B10', 'C10', 'D10', 'E10', 'F10', 'G10', 'H10', 'I10']
            row11 = ['A11']
            row12 = ['A12', 'B12', 'C12', 'D12', 'E12', 'F12', 'G12', 'H12', 'I12']
            row13 = ['B13', 'C13', 'D13', 'E13', 'F13', 'G13', 'H13', 'I13']
            all_rows = row1 + row2 + row3 + row4 + row5 + row6 + row7 + row8 + row9 + row10 + row11 + row12 + row13
            for i in range(len(data)):
                ws[all_rows[i]] = data[i]
                ws[all_rows[i]].font = bold
                ws[all_rows[i]].alignment = Alignment(horizontal='center', vertical='center')
            wb.save(path)

        except Exception:
            MyMessageBox(fr,'表格已打开，无法写入',wx.CAPTION)

    def getName(self):
        request.urlopen('http://202.197.144.243:'+self.port+'/xjInfoAction.do?oper=ydxx')
        self.resp = request.urlopen('http://202.197.144.243:'+self.port+'/userInfo.jsp').read()
        self.html = self.resp.decode('GBK')
        return self.html

    def info(self, html):
        # 用户名匹配
        r1 = r'<td >([\s\S]*?)</td>'
        # 匹配表格，这里会用到第二个表格
        r2 = r'<table width="100%" border="0" cellpadding="0" cellspacing="0" class="titleTop3">([\s\S]*?)</table>'
        # 匹配表格的电话
        r3 = r'<td >[\s\S]*?<input type="text" name="dh" value="([\s\S]*?)" tag="电话|0|s|40">[\s\S]*?</td>'
        # 匹配表格的地址
        r4 = r'<td>[\s\S]*?<input type="text" name="txdz" value="([\s\S]*?)" tag="通讯地址|0|s|100">[\s\S]*?</td>'
        # 匹配表格的Email
        r5 = r'<td>[\s\S]*?<input type="text" name="email" value="([\s\S]*?)" tag="电子邮件|0|e|100">[\s\S]*?</td>'

        # 找到的用户名
        name = re.findall(r1, html)
        # 找到的所有表
        table = re.findall(r2, html)
        # 找到的电话，这里匹配貌似有bug 。。。不管了
        phone = re.findall(r3, table[1])
        for i in phone:
            if i == '':
                phone.remove(i)
        # 找到的通讯地址
        address = re.findall(r4, table[1])
        for i in address:
            if i == '':
                address.remove(i)
        email = re.findall(r5, table[1])
        for j in email:
            if j == '':
                email.remove(j)

        # 放到字典里面
        info_dict = {
            'name': ''.join(name[0]).strip(),
            'phone': ''.join(phone[0]).strip(),
            'address': ''.join(address[0]).strip(),
            'email': ''.join(email[0]).strip()
        }
        return info_dict

    def exit(self):
        data ={
            'loginType' :' platformLogin'
        }
        data = parse.urlencode(data).encode('GBK')
        res = request.Request('http://202.197.144.243:'+self.port+'/logout.do', data, self.headers)
        resp =request.urlopen(res)

        if resp.read() is not '':
            return True


wildcard = "xlsx 文件 (*.xlsx)|*.xlsx|" \
           "All files (*.*)|*.*"

class MyMessageBox(wx.Dialog):
    def __init__(self,parent,message,style):
        super(MyMessageBox ,self).__init__(parent,title='提示',pos=wx.DefaultPosition,size=(200,140),style=style)
        panel =wx.Panel(self)
        self.text = wx.StaticText(panel,label=message,size=(100,20),pos=(35,15))
        self.btn =wx.Button(panel, wx.ID_OK, label="确定", size=(90, 30), pos=(90, 60))

def OnClicked(Event):
    if username.GetValue()is not '' and password.GetValue() is not '':
       try:
           usr = int(username.GetValue())
           psw = password.GetValue()
           port = rb.GetSelection()

           if port ==0:
               port = '8083'
           elif port ==1:
               port ='8084'
           else:
               port ='8085'
           try:
                urp =URP(usr,psw ,port)
                html = urp.login()

                try:
                    html1 = urp.getName()
                    info = urp.info(html1)
                    td_list = urp.getTable(html)
                    if td_list is not []:
                        r = wx.MessageBox('登录成功!\n\n你好，'+info.get('name'), '提示', wx.OK | wx.ICON_INFORMATION)
                        if r == 4:
                            r1 = wx.MessageBox('是否提取课表？', '询问', wx.YES_NO| wx.ICON_QUESTION)

                            if r1 == 2:
                                if sys.platform =='win32':
                                    path = 'C:\\Users\\'+getpass.getuser()+'\\Desktop'
                                else:
                                    path = '\\'

                                back = wx.FileDialog(fr, '选择文件', path, '课表.xlsx', wildcard=wildcard,
                                                     style=wx.FD_SAVE)
                                if back.ShowModal() == wx.ID_OK:
                                    urp.write_xlsx(back.GetPath(), td_list)
                                    wx.MessageBox('提取成功：' + back.GetPath(), '提示')
                                    back = urp.exit()
                                    if back:
                                        MyMessageBox(fr, '学号已退出！', wx.CAPTION).ShowModal()
                            else:
                                username.SetValue("")
                                password.SetValue("")
                                back = urp.exit()
                                if back:
                                    MyMessageBox(fr, '学号已退出！', wx.CAPTION).ShowModal()
                except Exception as e:

                    MyMessageBox(fr, '学号或密码错误！', wx.CAPTION).ShowModal()

           except Exception:
               MyMessageBox(fr,"网络异常",wx.CAPTION).ShowModal()

       except Exception as  e:
           MyMessageBox(fr, '学号有误！',wx.CAPTION).ShowModal()

    else:
        MyMessageBox(fr,'学号或密码为空！',wx.CAPTION).ShowModal()


app =wx.App()
fr = wx.Frame(None,size=(330,250),name='frame',style=wx.SIMPLE_BORDER|wx.CAPTION|wx.MINIMIZE_BOX|wx.CLOSE_BOX,pos=(600,300))
fr.CreateStatusBar()
fr.SetStatusText("   本程序由软工-某大神开发，仅适用湖北第二师范学院")
fr.SetTitle("一键提取课表")
fr.Show()
usrL = wx.StaticText(fr,-1,pos=(20,12))
usrL.SetLabel('学号')
username = wx.TextCtrl(fr,pos=(50,10))
username.SetMaxLength(10)
username.SetFocus()
pswL = wx.StaticText(fr,-1,pos=(20,55))
pswL.SetLabel('密码')
password = wx.TextCtrl(fr,pos=(50,50),style=wx.TE_PASSWORD)
password.SetMaxLength(20)
rb_list = ['8083','8084','8085']
rb = wx.RadioBox(fr,label='端口',pos =(20,100),choices=rb_list,majorDimension =1,style = wx.RA_SPECIFY_ROWS)
login =wx.Button(fr,pos=(200,10),size=(80,30))
login.SetLabel('登录')
login.Bind(wx.EVT_BUTTON,OnClicked)

app.MainLoop()

